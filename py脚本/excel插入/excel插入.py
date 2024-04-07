import openpyxl
from datetime import datetime
# 打开Excel文件
workbook = openpyxl.load_workbook('1.xlsx')

# 获取第一个工作表
sheet = workbook.active

# 获取总行数和总列数
total_rows = sheet.max_row
total_columns = sheet.max_column

# 循环读取每一行数据
for row_num in range(1, total_rows + 1):
    name = sheet.cell(row=row_num, column=1).value
    gender = sheet.cell(row=row_num, column=2).value
    ethnicity = sheet.cell(row=row_num, column=3).value
    ages = sheet.cell(row=row_num, column=4).value
    classes = sheet.cell(row=row_num, column=5).value
    xuehao = sheet.cell(row=row_num, column=6).value
    riqi1 = sheet.cell(row=row_num, column=7).value
    riqi2 = sheet.cell(row=row_num, column=8).value
    lianxiren = sheet.cell(row=row_num, column=9).value
    date_object = riqi1

    # 提取年份和月份
    year = date_object.year
    month = date_object.month
    day = date_object.day


    # 将年份和月份拼接成所需的格式
    formatted_date_bir1 = f"{year}年{month}月{day}日"

    date_object = riqi2

    # 提取年份和月份
    year = date_object.year
    month = date_object.month
    day = date_object.day
    formatted_date_bir2 = f"{year}年{month}月{day}日"
    # 提取年份和月份
    # # year = birth_date.year
    # # month = birth_date.month
    # # 将日期格式化为"YYYY年M月"的形式
    # age = sheet.cell(row=row_num, column=5).value
    # enrollment_date = sheet.cell(row=row_num, column=6).value
    # year = enrollment_date.year
    # month = enrollment_date.month
    #
    # # 将年份和月份拼接成所需的格式
    # formatted_date_enr = f"{year}年{month}月"
    # class_number = sheet.cell(row=row_num, column=7).value
    # student_id = sheet.cell(row=row_num, column=8).value
    # application_date = sheet.cell(row=row_num, column=9).value
    # year = application_date.year
    # month = application_date.month
    #
    # # 将年份和月份拼接成所需的格式
    # formatted_date_app = f"{year}年{month}月"

    # 按照需要处理每一行的数据，例如将数据插入到文本模板中或者进行其他操作
    formatted_info = "{}同志，{}，{}，{}岁，2022年9月进入计算机学院学习，现为{}班学生，学号：{}。{}同志于{}递交入党申请书，并于{}被确定为入党积极分子，培养联系人{}。党支部经过培养考察，该同志已基本具备党员条件，在听取党小组、培养联系人、党员和群众意见后，经支委会于2023年11月28日讨论同意并报计算机学院党委备案，2023年11月28日被列为发展对象。".format(
        name, gender, ethnicity, ages, classes, xuehao, name, formatted_date_bir1, name, formatted_date_bir2,lianxiren)

    # 输出格式化后的信息
    print(formatted_info)

# 关闭Excel文件
workbook.close()
